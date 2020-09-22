#Copy the content of sheet2 of all the excel files in a folder and paste it in a single excel file as different sheets.


import openpyxl as xl
import glob

#destination file name
path2 = 'Book1.xlsx'

#iterate through all the files in the folder
for file_name in glob.iglob('C:/Users/himanshu/Desktop/pythonn/*.xlsx', recursive=True):
	path1=file_name
	#skipping the destination file. Change path to the destination file
	if(path1=="C:/Users/himanshu/Desktop/pythonn\Book1.xlsx"):
		continue
	else:
		# print(file_name)
		wb1 = xl.load_workbook(filename=path1)
		ws1 = wb1.worksheets[1]#get sheet 2
		wb2 = xl.load_workbook(filename=path2)
		print("pasting sheet2 of ", path1," into ",path2,". What do you want to name it? ", end="")
		a=input()#enter the name you want to save it as
		ws2 = wb2.create_sheet(a)#create sheet with that name
		#copy all the data
		for row in ws1:
			for cell in row:
				ws2[cell.coordinate].value = cell.value
		#save file
		wb2.save(path2)